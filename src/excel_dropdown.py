import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def add_dropdowns_to_columns(input_file, output_file, sheet_name, dropdowns):
    """
    Adds dropdowns to specified columns in an Excel workbook.

    Args:
        input_file (str): Path to the input Excel file.
        output_file (str): Path to save the modified Excel file.
        sheet_name (str): Name of the sheet to add dropdowns to.
        dropdowns (dict): A dictionary where keys are column letters and values are lists of dropdown options.
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"The file {input_file} does not exist.")

    workbook = load_workbook(input_file)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"The sheet {sheet_name} does not exist in the workbook.")

    sheet = workbook[sheet_name]

    for column, dropdown_values in dropdowns.items():
        # Create a data validation object for the dropdown
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(dropdown_values)}"',
            allow_blank=True
        )

        # Apply the dropdown validation to the entire column
        dv.ranges.add(f"{column}1:{column}1048576")
        sheet.add_data_validation(dv)

    # Save the modified workbook to the output file
    workbook.save(output_file)
    print(f" ✅  Dropdowns added successfully. Saved to {output_file}")
