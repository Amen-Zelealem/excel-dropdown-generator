import os
from src.excel_dropdown import add_dropdowns_to_columns
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# Input and output paths
input_file = "data/Record1.xlsx"  # Replace with your actual input file name
os.makedirs("output", exist_ok=True)
output_file = "output/output_workbook1.xlsx"  # Save to the output folder
sheet_name = "Sheet1"  # Replace with the actual sheet name

# Define dropdowns for other columns
dropdowns = {
    "A": ["Abebe", "Kebede", "Option3"],
    "B": ["Yes", "No", "Maybe"],
    "C": ["Red", "Green", "Blue"],
    "D": ["2021", "2022", "2023"]
}

# Call the function to add dropdowns
add_dropdowns_to_columns(input_file, output_file, sheet_name, dropdowns)

# Add instructions to Column E for adding images
workbook = load_workbook(output_file)
sheet = workbook[sheet_name]

# Add header text for Column E
sheet["E1"] = "Insert Image Here"
sheet["E1"].font = Font(bold=True, color="FF0000")  # Red and bold font
sheet["E1"].alignment = Alignment(horizontal="center")

# Add placeholder text to rows in Column E
for row in range(2, sheet.max_row + 1):
    sheet[f"E{row}"] = "Right-click and insert image"
    sheet[f"E{row}"].alignment = Alignment(horizontal="center")

# Save the updated workbook
workbook.save(output_file)
print(f"Workbook prepared for adding images. Saved to {output_file}")
